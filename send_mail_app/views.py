from django.shortcuts import render,HttpResponse,redirect,HttpResponseRedirect
from send_mail_app.tasks import sending_mail
# from send_mail_app.tasks import send_mail_function
# from send_mail_app.tasks import send_mail_with_atachments
from send_mail_app.tasks import send_mail_with_attachments
from django.conf import settings
from .models import Employee,Customer,EmailHistory
# Create your views here.
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook,load_workbook
from django.core.mail import EmailMessage
# from send_mail_app.tasks import send_b_mail
from send_mail_app.tasks import doc_to_pdf
import json
import uuid
import json
import datetime
from .forms import NewUserForm
from django.contrib.auth import login,authenticate
from django.contrib.auth import login as auth_login    
from django.contrib.auth.forms import AuthenticationForm 
from django.contrib import messages    

from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django_celery_beat.models import PeriodicTask, IntervalSchedule, CrontabSchedule
from django.core.files.storage import FileSystemStorage
from docx2pdf import convert
from celery.result import AsyncResult

import os

def send_mail_to_gmail(request):
    send_mail_function.delay()
    return HttpResponse("mail send check it")

# def send_mail_to_gmail(request):
    
#     subject="Periodic Task"
#     message="Happy Teachers day to everyone"                       
#     # from_email=settings.EMAIL_HOST_USER,
#     recipient_list=['raiyanar0@gmail.com']
#     file_path = f"{settings.BASE_DIR}/mail.xlsx"
#     send_mail_with_atachments(subject,message,recipient_list,file_path)
#     return HttpResponse("mail send check it")

def send_xlsx(request):
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                       
    subject="Periodic Task"
    message="Hello this is a simple xlsx attachment"                       
   
    user=request.user
    print("user")    
    
    print(user)
        
    recipient_list=['raiyanar0@gmail.com']
    file_path = f"{settings.BASE_DIR}/Employee-2023-09-06.xlsx"
    send_mail_with_attachments.delay(subject,message,recipient_list,file_path)
    return HttpResponse("attachment send check it")

    
# def send_mail_to_gmail(request):
#     send_mail_with_atachments(subject,message,recipient_list,file_path)
#     return HttpResponse("mail send check it")

def send_email(request):
    email1=("subject3","message3","raiyanar0@gmail.com",['raiyanar0@gmail.com'])
    email2=("subject4","message4","raiyyanabdurrehman@gmail.com",['raiyyanabdurrehman@gmail.com'])    
    messages=(email1,email2)
    return messages
    sending_mail(messages)
    return HttpResponse("email send successfully")

import pandas as pd
from django.http import JsonResponse
def export_data_xls(request):
    objs=Employee.objects.all()
    print("objs")
    print(objs)
    pd.DataFrame(objs).to_excel('output.xlxs')
    return JsonResponse({
        'status':200
    })
        
def register(request):
  if request.method=="POST":
    form=NewUserForm(request.POST)
    if form.is_valid():
      user=form.save()
      messages.success(request,"registration Successful")
      return redirect("/login")
    messages.error(request,"Unsuccessful")      
  else:
    
   form=NewUserForm()
  return render(request,'register.html',context={"register_form":form})


def login_request(request): 
  if request.method=="POST":
    print("Form was submitted")
    form = AuthenticationForm(request, data=request.POST)
    if form.is_valid():
      print("valid")
      print(form.is_valid())
      username = form.request.POST.get('username')  
      password = form.request.POST.get('password')
      user = authenticate(username=username, password=password)
      # user=form.save()
      if user:
        login(request,user)
        messages.success(request,"login successful")
        print("asddas")
        return redirect("logout")
           
    messages.error(request,"Unsuccessful")
    print(form.errors)    
     
  else:
    form=AuthenticationForm()
  return render(request,'login.html',context={"login_form":form})


def logout(request):
  
  users=get_user_model().objects.all()
  reciepient_list=[user.email for user in users]
  print("reciepient_list")
  print(reciepient_list)
  
  return render(request,"logout.html",{"users":users,"reciepient_list":reciepient_list})

import openpyxl
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

# def download_data(request):
    
#     # customer=Customer.objects.get(user=request.user)
#     response =HttpResponse(content_type='application/ms-excel')
#     response['Content-Disposition'] ='attachment;filename="customers.xlxs"'
#     wb=Workbook()
#     ws=wb.active
#     print("ws")
#     print(ws)
#     # ws.title="Customers"
    
#     headers=["Name","Location","City"]
#     ws.append(headers)
    
#     # customer=Customer.objects.get(user=request.user)
#     customers=Customer.objects.all()
#     print("customers")
#     print(customers)
    
#     for customer in customers:
#         print("Customer name")
#         # print(customer.name)
#         customer_name=str(customer.name)
#         customer_location=str(customer.location)
#         customer_city=str(customer.city)
        
#         # print(type(customer.name))
#         # print(type(customer_name))
        
#         # print(customer.location)
#         # print(customer.city)
#         # print(type(customer.location))
        
#         ws.append([customer_name,customer_location,customer_city])
        
#     wb.save(response)
#     return response
    
def download_data(request):
    
    # customer=Customer.objects.get(user=request.user)
    response =HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] ='attachment;filename="customers.xlxs"'
    wb=Workbook()
    ws=wb.active
    print("ws")
    print(ws)
    # ws.title="Customers"
    
    headers=["Name","Location","City"]
    ws.append(headers)
    
    customers=Customer.objects.get(name=request.user)
    # customers=Customer.objects.all()
    print("customers")
    print(customers)
    print("Customer name")
    # print(customer.name)
    print(type(customers.name))
    print(customers.name.email)
    password=request.user.password
    print("password")
    print(password)
    firstname=request.user.first_name
    print("firstname") 
    print(firstname) 
    
    # customer_name=str(customer.name)
    # customer_location=str(customer.location)
    # customer_city=str(customer.city)
        
        # print(type(customer.name))
        # print(type(customer_name))
        
        # print(customer.location)
        # print(customer.city)
        # print(type(customer.location))
        
        
    # ws.append([customer_name,customer_location,customer_city])
    ws.append([customers.name.username,customers.location,customers.city])
        
    wb.save(response)
    # print("response")
    # print(response)
    # for x in response:
    #     print(x)
    return response

import shutil
def download_data1(request):
    
    # customer=Customer.objects.get(user=request.user)
    response =HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] ='attachment;filename="customers.xlxs"'
    wb=Workbook()
    ws=wb.active
    print("ws")
    print(ws)
    
    # original=load_workbook('/send_mail/static/mail.xlsx')
    
    
    original = ws['mail']
    
    # target=
    # shutil.copyfile(original,target)
    
    
    print("original")
    print(original)
    
    # ws.title="Customers"
    
    headers=["Name","Location","City"]
    ws.append(headers)
    
    customers=Customer.objects.get(name=request.user)
    # customers=Customer.objects.all()
    print("customers")
    print(customers)
    print("Customer name")
    # print(customer.name)
    print(type(customers.name))
    print(customers.name.email)
    
    ws.append([customers.name.username,customers.location,customers.city])    
    wb.save(response)
    
    print("response")
    print(response)
    # send_mail_with_atachments("subject","message",["raiyanar0@gmail.com"],response)
    file_path = f"{settings.BASE_DIR}/mail.xlsx"
    print("file_path")
    print(file_path)
    file_split=file_path.split("send_mail")[1]
    new_url= f"http://127.0.0.1:8000/static{file_split}"
    print(file_split)
    print(new_url)
    target_file=excel_writer(request,new_url,new_url)
    return redirect( target_file)
  
    
# def excel_writer(request,priginal,target):
#   return target


def attach_file(request):
    
    subject="Excel data"
    message="Hello this is a exported data"                       
    user=request.user
    email=request.user.email
    
    print("user")      
    print(user)
    print("email")
    print(email)
       
    
    recipient_list=[email]
    print("recipient list")
    print(recipient_list)
    
    file_path = f"{settings.BASE_DIR}/mail.xlsx"
    
    # file_path = str(settings.BASE_DIR)+'/mail.xlsx'

    # target_file=excel_writer(request,file_path,file_path)
    send_mail_with_atachments(subject,message,recipient_list,file_path)
    
    return HttpResponse("attachment send check it")


from django.templatetags.static import static

import shutil
import openpyxl
import pandas as pd


def attach_file1(request):
    original='C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/static/original.xlsx'
    print("original")
    print(original)
    # original = 'C:\\Users\\ABU DUJANA ANSARI\Desktop\\original.xlsx'
    target = 'C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/static/target.xlsx'
    print("target")
    print(target)
   
    shutil.copyfile(original, target)

    wb = openpyxl.load_workbook(target)
    print("wb")
    print(wb)
    
    ws = wb.active
    print("ws sheet")
    print(ws)

    customers=Customer.objects.get(name=request.user)

    print("customers")
    print(customers)   

    data=[{'name': customers.name.username,
    'location': customers.location, 
    'city': customers.city}]
    
    # for obj in customers:
    #    data.append({
    #       'name': obj.name.username,
    #       'location': obj.location, 
    #       'city': obj.city
    #    })
       
    print("data")
    print(data)

    pd.DataFrame(data).to_excel(target)

    file_path1='C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/static/original.xlsx'
    print("file_path1")
    print(file_path1)
    file_split1=file_path1.split("static")[1]
    print("file_split1")
    print(file_split1)

    old_url= f"http://127.0.0.1:8000/static{file_split1}"
    print("old_url")
    print(old_url)

    file_path2='C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/static/target.xlsx'
    print("file_path2")
    print(file_path2)
    file_split2=file_path2.split("static")[1]
    print("file_split2")
    print(file_split2)

    new_url= f"http://127.0.0.1:8000/static{file_split2}"
    print("new_url")
    print(new_url)
    
    user=request.user
    email=request.user.email
   
    print("user")      
    print(user)
    print("email")
    print(email)
  
    subject="Excel data"
    message=f"Hello this is a exported data:\n {new_url},Your credentials are: Username:{request.user},Password:{request.user.password}"    
    
    print("message")
    print(message)
    recipient_list=[email]
    print("recipient list")
    print(recipient_list)
    
    # file_paths=[file_path1,file_path2]
    # print("file_paths")
    # print(file_paths)
    
    target_file=excel_writer(request,old_url,new_url)

    send_mail_with_attachments.delay(subject,message,recipient_list,file_path2)

    # return redirect( target_file)
    return HttpResponse("Email send successfully")

def excel_writer(request,old_url,new_url):
  return new_url
  # return HttpResponse("ok")




def file_convert(request):
  if request.method =="POST":
    files=request.FILES['files']
    print("files")
    print(files)

    fs=FileSystemStorage()
    print("fs")
    print(fs)
    namefile=fs.save(files.name,files)
    print("namefile")
    print(namefile)
    
    uploaded_file=fs.url(namefile)
    print("uploaded_file")
    print(uploaded_file)
    
    data=doc_to_pdf.delay(files.name)
    print("data")
    print(data)
    
    # return HttpResponseRedirect(data.task_id)
    return HttpResponse("file converted to pdf")
    
    # convert('static/' + files.name)    
  return render(request,'file.html')


def checkstatus(request,task_id):
  print("TASK_ID")
  print(task_id)
  res=AsyncResult(task_id)  
  print(res.ready())
  task_status=res.ready()
  print(task_status)
  context={'task_status':task_status}  
  return render(request,'checkstatus.html',context)



def mailing(request):
  response={}
  if request.method=="POST":
    try:      
      email_id=request.POST['email_id']    
      full_date=request.POST['date']
      timing=request.POST['timing']
      refresh_period = request.POST.get('refresh_period', None)
      print("Refresh Period Selected:", refresh_period)
      
      print("full_date")
      get_date=full_date
      if get_date:
        full_date=full_date.split("-")
        print(full_date)
        year=full_date[0]
        print("year")
        print(year)
        print("month")
        month=full_date[1]
        print(month)
        date=full_date[2]
        
        print("date")
        print(date)
     
      if timing:
         
        print("timing")
        timing=timing.split(":")
        print(timing)
        hour=timing[0]
        minutes=timing[1]
        print("hour")
        print(hour)
        print("minutes")
        print(minutes)
      
      if not email_id:
        return HttpResponse("Please Enter email_id")
        
    except Exception as e:
      response['error'] = True 
      response['message'] = str(e)
      print("Exception occurred:", e)
            
    file_path2= "C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/mail.xlsx"   
    user=request.user
    email=request.user.email
    host_user_email=settings.EMAIL_HOST_USER

    import re    
    def check(s):
      pat = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
      if re.match(pat,s):
        return True
      else:
       return False
     
    if check(host_user_email) and check(email):
     status="Success"
    else:
     status="Failed"
  
    model=EmailHistory(email_from=host_user_email,email_to=email,status=status)
    model.save()    
    
    print("user")      
    print(user)
    print("email")
    print(email)
    subject="Excel data"
    message=f"Hello this is a exported data,Your credentials are: Username:{request.user},Password:{request.user.password},FirstName:{request.user.first_name},LastName:{request.user.last_name}"    
    print("message")
    print(message)
    recipient_list=[email]
    print("recipient list")
    print(recipient_list)
      
    
    if get_date and timing:      
        chon_schedule = CrontabSchedule.objects.create(
          minute=minutes, 
          hour=hour,          
          day_of_month=date, 
          month_of_year=month,
          day_of_week='*', 
      )
        args = (subject, message, recipient_list, file_path2)
        create_schedule = PeriodicTask.objects.create(
          name="xyz" + str(uuid.uuid4()),
          crontab=chon_schedule,
          task='send_mail_app.tasks.send_mail_with_attachments',
          args=json.dumps(args)
      )      
        return HttpResponse("Email schedule successfully")      
    else:
      send_mail_with_attachments.delay(subject, message, recipient_list, file_path2)
      return HttpResponse("Email sent successfully")



def mailing_all(request):      
  
  file_path2= "C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/mail.xlsx"   
  subject="Your Email Details"
  users=get_user_model().objects.all()
  
  import re
  
  def check(s):
    pat = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
    if re.match(pat,s):
      return True
    else:
       return False
     
  host_user_email=settings.EMAIL_HOST_USER        
  users=get_user_model().objects.all()
  
  for user in users:
    print("mailing")       
    to_email=user.email
    username=user.username
    password=user.password
    firstname=user.first_name
    lastname=user.last_name
    
    if check(host_user_email) and check(to_email):
       status="Success"
    else:
       status="Failed"
    
    model=EmailHistory(email_from=host_user_email,email_to=to_email,status=status)
    model.save()

  message=f"Hello ,Your credentials are: Username:{username},Password:{password},firstname:{firstname},lastname:{lastname}"
  send_mail_with_attachments.delay(subject,message,[to_email],file_path2)
  
  return HttpResponse("Email send successfully")


# def mailing_all(request):
  
#   users=get_user_model().objects.all()
  
#   selected_email=request.POST.get('email_all')
#   if selected_email=='reciepient_list':
#     reciepients =[user.email for user in users]
#   else:
#     reciepients=[selected_email]
  
#   file_path2= "C:/Users/Rahul - Arivani/Desktop/celery messages/send_mail/mail.xlsx"   
#   subject="Your Email Details"
#   for reciepient in reciepients:
#     user=get_user_model().objects.get(email=reciepient)
    
#     print("mailing")       
#     to_email=user.email
#     username=user.username
#     password=user.password
#     firstname=user.first_name
#     lastname=user.last_name
  
#     message=f"Hello ,Your credentials are: Username:{username},Password:{password},firstname:{firstname},lastname:{lastname}"
    
#     send_mail_with_attachments.delay(subject,message,[to_email],file_path2)
    
#   return HttpResponse("Email send successfully")


def email_history(request):
  email_history=EmailHistory.objects.all().order_by('-dateTime')
  return render(request,'email_history.html',{'email_history':email_history})
 