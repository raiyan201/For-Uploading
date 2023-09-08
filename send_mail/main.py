import openpyxl

from openpyxl import Workbook,load_workbook
book=load_workbook('send_mail/media/employee.xlsx')

sheet=book.active
sheet['B2'].value='JILL'
print(sheet['B2'].value)

# print(sheet.cell(row=3,column=2).value)

book.save('send_mail/media/employee4.xlsx')

                                                                            